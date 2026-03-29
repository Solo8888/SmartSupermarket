# Footfall analytics routes
# Provides API endpoints for footfall analysis

from fastapi import APIRouter, HTTPException, Query, Depends
from fastapi.responses import Response, StreamingResponse
from typing import Optional, IO
import io
import csv
import json
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.charts.axes import XCategoryAxis
from reportlab.graphics.charts.textlabels import Label
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from .schemas import (
    TimeDistributionResponse, TimeDistributionQuery, ExportRequest,
    WeekComparisonResponse, WeekendWeekdayResponse, ForecastResponse
)
from .service import TimeDistributionService, ExportService, ComparisonService, ForecastService
from core.permitions import require_role

footfall_router = APIRouter(
    prefix="",
    tags=["footfall"],
    responses={404: {"description": "Not found"}},
)

time_distribution_service = TimeDistributionService()
export_service = ExportService()
comparison_service = ComparisonService()
forecast_service = ForecastService()


@footfall_router.get("/time-distribution", response_model=TimeDistributionResponse)
async def get_time_distribution(
    start_date: str = Query(..., description="开始日期 (YYYY-MM-DD)"),
    end_date: str = Query(..., description="结束日期 (YYYY-MM-DD)"),
    store_id: Optional[str] = Query(None, description="门店ID（可选）"),
    current_user = Depends(require_role(["operations_manager"], mode="in"))
):
    """获取时段客流分布
    
    - **start_date**: 开始日期 (YYYY-MM-DD)
    - **end_date**: 结束日期 (YYYY-MM-DD)
    - **store_id**: 门店ID（可选）
    """
    try:
        from datetime import datetime
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
        
        if start > end:
            raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
        
        data = time_distribution_service.get_time_distribution(start_date, end_date, store_id)
        
        response = TimeDistributionResponse(
            data=data
        )
        
        return response
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取时段客流分布失败: {str(e)}")


@footfall_router.post("/export")
async def export_footfall_report(
    request: ExportRequest,
    current_user = Depends(require_role(["operations_manager"], mode="in"))
):
    """导出客流分析报告
    
    - **start_date**: 开始日期 (YYYY-MM-DD)
    - **end_date**: 结束日期 (YYYY-MM-DD)
    - **store_id**: 门店ID（可选）
    - **format**: 导出格式 (pdf/excel)
    """
    try:
        from datetime import datetime
        start = datetime.strptime(request.start_date, "%Y-%m-%d")
        end = datetime.strptime(request.end_date, "%Y-%m-%d")
        
        if start > end:
            raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
        
        if request.format not in ["pdf", "excel"]:
            raise HTTPException(status_code=400, detail="不支持的导出格式，支持的格式：pdf, excel")
        
        raw_data, hourly_counts = export_service.get_footfall_data(
            request.start_date, 
            request.end_date, 
            request.store_id
        )
        
        if request.format == "excel":
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "时段客流分布"
            
            # 添加标题
            ws['A1'] = "客流分析报告"
            ws['A3'] = f"时间范围：{request.start_date} 至 {request.end_date}"
            if request.store_id:
                ws['A4'] = f"门店ID：{request.store_id}"
            
            # 添加表头
            ws['A6'] = "小时"
            ws['B6'] = "客流量"
            
            # 添加数据
            row = 7
            hours = []
            counts = []
            for hour, count in sorted(hourly_counts.items()):
                ws[f'A{row}'] = hour
                ws[f'B{row}'] = count
                hours.append(hour)
                counts.append(count)
                row += 1
            
            # 添加图表
            chart = LineChart()
            chart.title = "时段客流分布"
            chart.style = 13
            chart.x_axis.title = "小时"
            chart.y_axis.title = "客流量"
            
            # 设置数据范围
            data = Reference(ws, min_col=2, min_row=6, max_row=row-1)
            categories = Reference(ws, min_col=1, min_row=7, max_row=row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # 将图表添加到工作表
            ws.add_chart(chart, "D6")
            
            # 调整列宽
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['D'].width = 30
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=footfall_report_{request.start_date}_{request.end_date}.xlsx"
                }
            )
        elif request.format == "pdf":
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            
            try:
                pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
                chinese_font = 'STSong-Light'
            except:
                try:
                    pdfmetrics.registerFont(UnicodeCIDFont('AdobeSongStd-Light'))
                    chinese_font = 'AdobeSongStd-Light'
                except:
                    chinese_font = 'Helvetica'
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=chinese_font,
                fontSize=24,
                spaceAfter=30,
                alignment=1
            )
            
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles["Normal"],
                fontName=chinese_font,
                fontSize=12
            )
            
            elements = []
            
            elements.append(Paragraph("客流分析报告", title_style))
            elements.append(Spacer(1, 20))
            
            elements.append(Paragraph(f"<b>时间范围：</b>{request.start_date} 至 {request.end_date}", info_style))
            if request.store_id:
                elements.append(Paragraph(f"<b>门店ID：</b>{request.store_id}", info_style))
            elements.append(Spacer(1, 20))
            
            # 添加图表
            elements.append(Paragraph("<b>时段客流分布图表</b>", info_style))
            elements.append(Spacer(1, 10))
            
            # 创建柱状图
            from reportlab.graphics.charts.barcharts import VerticalBarChart
            from reportlab.graphics.shapes import Drawing
            
            drawing = Drawing(500, 300)
            
            # 准备数据
            hours = []
            counts = []
            for hour, count in sorted(hourly_counts.items()):
                hours.append(hour)
                counts.append(count)
            
            # 创建柱状图
            chart = VerticalBarChart()
            chart.x = 60
            chart.y = 50
            chart.width = 420
            chart.height = 200
            
            # 设置数据
            chart.data = [counts]
            # 每2小时显示一个标签
            filtered_hours = [hour if i % 2 == 0 else '' for i, hour in enumerate(hours)]
            chart.categoryAxis.categoryNames = filtered_hours
            
            # 设置样式
            chart.bars[0].fillColor = colors.blue
            
            # 设置标签
            chart.categoryAxis.labels.fontName = chinese_font
            chart.categoryAxis.labels.fontSize = 9
            chart.valueAxis.labels.fontName = chinese_font
            chart.valueAxis.labels.fontSize = 8
            
            # 添加图表到drawing
            drawing.add(chart)
            
            elements.append(drawing)
            elements.append(Spacer(1, 30))
            
            # 添加数据表格
            elements.append(Paragraph("<b>时段客流分布数据</b>", info_style))
            elements.append(Spacer(1, 10))
            
            table_data = [["小时", "客流量"]]
            for hour, count in sorted(hourly_counts.items()):
                table_data.append([hour, str(count)])
            
            table = Table(table_data, colWidths=[200, 200])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), chinese_font),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), chinese_font),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            
            doc.build(elements)
            
            pdf_content = buffer.getvalue()
            buffer.close()
            
            return Response(
                content=pdf_content,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename=footfall_report_{request.start_date}_{request.end_date}.pdf"
                }
            )
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Export error: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"导出客流分析报告失败: {str(e)}")


@footfall_router.get("/week-comparison", response_model=WeekComparisonResponse)
async def get_week_comparison(
    store_id: Optional[str] = Query(None, description="门店ID（可选）"),
    current_user = Depends(require_role(["operations_manager"], mode="in"))
):
    """获取本周与上周每小时客流对比
    
    - **store_id**: 门店ID（可选）
    """
    try:
        data = comparison_service.get_week_comparison(store_id)
        return WeekComparisonResponse(data=data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取周对比数据失败: {str(e)}")


@footfall_router.get("/weekend-weekday-comparison", response_model=WeekendWeekdayResponse)
async def get_weekend_weekday_comparison(
    store_id: Optional[str] = Query(None, description="门店ID（可选）"),
    current_user = Depends(require_role(["operations_manager"], mode="in"))
):
    """获取工作日与周末每小时客流对比
    
    - **store_id**: 门店ID（可选）
    """
    try:
        data = comparison_service.get_weekend_weekday_comparison(store_id)
        return WeekendWeekdayResponse(data=data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取周末工作日对比数据失败: {str(e)}")


@footfall_router.get("/forecast", response_model=ForecastResponse)
async def get_forecast_data(
    date: str = Query(..., description="预测日期 (YYYY-MM-DD)"),
    store_id: Optional[str] = Query(None, description="门店ID（可选）"),
    current_user = Depends(require_role(["operations_manager"], mode="in"))
):
    """获取预测客流数据
    
    基于历史数据生成预测值
    
    - **date**: 预测日期 (YYYY-MM-DD)
    - **store_id**: 门店ID（可选）
    """
    try:
        from datetime import datetime
        datetime.strptime(date, "%Y-%m-%d")
        
        data = forecast_service.get_forecast_data(date, store_id)
        return ForecastResponse(data=data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"日期格式错误: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取预测数据失败: {str(e)}")
